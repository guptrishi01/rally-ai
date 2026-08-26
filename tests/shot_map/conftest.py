from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from swingvision_import.config import ImportConfig
from tests.swingvision_import.conftest import add_settings_and_shots_sheets


@pytest.fixture
def synthetic_non_pro_xlsx(tmp_path: Path) -> Path:
    """Same confirmed-real shape as tests/swingvision_import/conftest.py's
    fixture of the same name: Points/Games empty, Shots populated - the
    only path that carries shot-to-point provenance."""
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )
    shot_rows = [
        [1, 1, "Test Player", "first_serve", "Serve", "In"],
        [1, 2, "Test Opponent", "first_return", "Backhand", "Net"],
        [2, 1, "Test Player", "first_serve", "Serve", "In"],
        [2, 2, "Test Opponent", "first_return", "Backhand", "In"],
        [2, 3, "Test Player", "serve_plus_one", "Forehand", "Out"],
    ]
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows)
    path = tmp_path / "synthetic_non_pro_export.xlsx"
    workbook.save(path)
    return path


@pytest.fixture
def synthetic_xlsx(tmp_path: Path) -> Path:
    """A direct Points-sheet parse case (real rows, unlike the empty-Points
    fixture above) - its points get no source_point_number at all, so
    there's no shot<->point provenance for build_shot_embeddings to use."""
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set #", "Winner", "Games Won", "Games Lost"])
    sets_sheet.append([1, "host", 6, 4])
    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Set #", "Game #", "Server", "Winner"])
    games_sheet.append([1, 1, "host", "host"])
    points_sheet = workbook.create_sheet("Points")
    points_sheet.append(
        [
            "Set #", "Game #", "Point #", "Server", "Winner",
            "1st Serve In", "2nd Serve In", "Shot Type",
        ]
    )
    points_sheet.append([1, 1, 1, "host", "host", True, None, "ace"])
    add_settings_and_shots_sheets(workbook)
    path = tmp_path / "synthetic_export.xlsx"
    workbook.save(path)
    return path


@pytest.fixture
def import_config(tmp_path: Path) -> ImportConfig:
    return ImportConfig(pending_dir=tmp_path / "pending", db_path=tmp_path / "rallyai.db")
