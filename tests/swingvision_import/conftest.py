from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from swingvision_import.config import ImportConfig


def add_settings_and_shots_sheets(
    workbook: Workbook,
    *,
    host_name: str = "Test Player",
    guest_name: str = "Test Opponent",
    shot_rows: list[list] | None = None,
) -> None:
    """Appends minimal Settings/Shots sheets to a hand-built test workbook.

    Every real export has these two sheets, and parse() now always reads
    both — a workbook missing either raises a KeyError unrelated to
    whatever the test actually cares about. Most tests just need these
    present and structurally valid, not populated with meaningful data.

    Args:
        workbook: The openpyxl Workbook being built.
        host_name: Value for Settings' "Host Team" column.
        guest_name: Value for Settings' "Guest Team" column.
        shot_rows: Optional data rows for the Shots sheet, each
            [point, shot, player, type, stroke, result]. Defaults to none
            (header only), which parses to an empty shot list.
    """
    settings_sheet = workbook.create_sheet("Settings")
    settings_sheet.append(["Host Team", "Guest Team"])
    settings_sheet.append([host_name, guest_name])

    shots_sheet = workbook.create_sheet("Shots")
    shots_sheet.append(["Point", "Shot", "Player", "Type", "Stroke", "Result"])
    for row in shot_rows or []:
        shots_sheet.append(row)


@pytest.fixture
def synthetic_xlsx(tmp_path: Path) -> Path:
    """A small hand-built .xlsx standing in for a real, Pro-tier SwingVision
    export where Points/Games are actually populated.

    Column *names* are now confirmed against two real (non-Pro) exports —
    see raw.py/config.py. Points-sheet *data* is still unverified (no Pro
    export has been seen with real rows), so this fixture's Points sheet
    still stands in for a hypothetical direct-parse scenario, exercised by
    the older transform.py path. See synthetic_non_pro_xlsx for the
    confirmed-real shape (Points empty, Shots populated).
    """
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set #", "Winner", "Games Won", "Games Lost"])
    sets_sheet.append([1, "host", 6, 4])

    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Set #", "Game #", "Server", "Winner"])
    games_sheet.append([1, 1, "host", "host"])
    games_sheet.append([1, 2, "guest", "guest"])

    points_sheet = workbook.create_sheet("Points")
    points_sheet.append(
        ["Set #", "Game #", "Point #", "Server", "Winner",
         "1st Serve In", "2nd Serve In", "Shot Type"]
    )
    points_sheet.append([1, 1, 1, "host", "host", True, None, "ace"])
    points_sheet.append([1, 1, 2, "host", "host", False, True, "winner"])
    # Guest serves, host commits an unforced error - guest wins the point,
    # consistent with the point/point_end_type CHECK constraint in
    # data/schema.sql (an unforced_error always means point_won=False).
    points_sheet.append([1, 2, 1, "guest", "guest", True, None, "unforced_error"])

    add_settings_and_shots_sheets(workbook)

    path = tmp_path / "synthetic_export.xlsx"
    workbook.save(path)
    return path


@pytest.fixture
def synthetic_non_pro_xlsx(tmp_path: Path) -> Path:
    """A hand-built .xlsx matching the confirmed real (non-Pro) shape: Sets
    has real summary rows, Points/Games are empty, Shots is fully
    populated. Exercises the pipeline's fallback-to-reconstruction routing.
    """
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])

    games_sheet = workbook.create_sheet("Games")
    games_sheet.append(["Game", "Set", "Server", "Game Winner"])

    points_sheet = workbook.create_sheet("Points")
    points_sheet.append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )

    # Point 1: host aces. Point 2 is a gap (no shots at all), matching the
    # real ~15 missing point numbers seen in the actual export. Point 3:
    # host serves, guest nets a return - host wins.
    shot_rows = [
        [1, 1, "Test Player", "first_serve", "Serve", "In"],
        [3, 1, "Test Player", "first_serve", "Serve", "In"],
        [3, 2, "Test Opponent", "first_return", "Backhand", "Net"],
    ]
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows)

    path = tmp_path / "synthetic_non_pro_export.xlsx"
    workbook.save(path)
    return path


@pytest.fixture
def import_config(tmp_path: Path) -> ImportConfig:
    return ImportConfig(
        pending_dir=tmp_path / "pending",
        db_path=tmp_path / "rallyai.db",
    )
