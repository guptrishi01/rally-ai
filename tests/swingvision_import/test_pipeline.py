from __future__ import annotations

import json
import logging
import sqlite3

import pytest
from openpyxl import Workbook

from swingvision_import.load import UnresolvedReviewError
from swingvision_import.pipeline import SwingVisionImportPipeline
from swingvision_import.records import MatchRecord
from swingvision_import.review import ConfirmationError, load_pending, save_pending
from tests.ai.conftest import FakeMessage, FakeTextBlock
from tests.swingvision_import.conftest import add_settings_and_shots_sheets


def _clear_review_flags(pipeline, json_path):
    record = load_pending(json_path)
    for set_record in record.sets:
        for point in set_record.points:
            point.needs_review = False
    save_pending(record, pipeline.config.pending_dir)


def test_ingest_writes_pending_json_without_touching_sql(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")

    assert json_path.exists()
    assert not import_config.db_path.exists()


def test_finalize_rejects_until_flags_resolved(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")

    with pytest.raises(UnresolvedReviewError):
        pipeline.finalize(json_path)


def test_finalize_loads_once_flags_are_cleared(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    _clear_review_flags(pipeline, json_path)

    match_id = pipeline.finalize(json_path)
    assert match_id == 1


def test_rerunning_finalize_on_same_match_is_rejected_not_duplicated(synthetic_xlsx, import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    _clear_review_flags(pipeline, json_path)

    pipeline.finalize(json_path)
    with pytest.raises(ValueError):
        pipeline.finalize(json_path)


def test_match_overrides_survive_the_full_ingest_then_finalize_round_trip(
    synthetic_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_xlsx,
        date="2026-08-06",
        opponent="Alex",
        result="W",
        energy_rating=5,
        pros="Aggressive on return games",
    )
    _clear_review_flags(pipeline, json_path)
    pipeline.finalize(json_path)

    connection = sqlite3.connect(import_config.db_path)
    row = connection.execute("SELECT energy_rating, pros FROM match").fetchone()
    connection.close()
    assert row == (5, "Aggressive on return games")


def test_reingesting_the_same_match_overwrites_the_pending_file_at_the_same_path(
    synthetic_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    first_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    second_path = pipeline.ingest(
        synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W", pros="Revised notes"
    )

    assert first_path == second_path
    assert list(import_config.pending_dir.glob("*.json")) == [first_path]


def test_ingest_falls_back_to_shot_reconstruction_when_points_sheet_is_empty(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)

    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert len(all_points) == 2
    assert all(p.needs_review for p in all_points)
    assert all(p.source_point_number is not None for p in all_points)
    # Point 1 was a single In serve shot by the host - an ace.
    assert any(p.point_end_type == "ace" for p in all_points)


def test_ingest_populates_import_notes_with_quality_check_and_gap_findings(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)

    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx,
        date="2026-08-06",
        opponent="Alex",
        result="W",
        first_server_by_set={1: "opponent"},  # deliberately wrong ground truth
        tracked_identity="Someone Else",  # deliberately wrong identity
    )
    record = load_pending(json_path)

    notes = " | ".join(record.import_notes)
    assert "had no shot data" in notes  # point 2 is a gap
    assert "reversed" in notes  # serve-order mismatch
    assert "Someone Else" in notes  # identity mismatch
    assert "does not match" in notes  # reconstructed score vs Sets-sheet summary


def test_ingest_skips_serve_order_and_identity_checks_when_not_supplied(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)

    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)

    notes = " | ".join(record.import_notes)
    assert "reversed" not in notes
    assert "identified yourself" not in notes


def test_ingest_matching_identity_produces_no_identity_note(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)

    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx,
        date="2026-08-06",
        opponent="Alex",
        result="W",
        tracked_identity="Test Player",
    )
    record = load_pending(json_path)

    assert "identified yourself" not in " | ".join(record.import_notes)


def test_ingest_reports_excluded_non_match_points_in_import_notes(tmp_path, import_config):
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 0, 0, "host"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )
    shot_rows = [
        [1, 1, "Test Player", "first_serve", "Serve", "In"],
        [1, 2, "Test Opponent", "first_return", "Backhand", "Net"],
        # Point 2: a fed ball, not real rallying - should be excluded, not
        # counted as a real point.
        [2, 0, "Test Player", "none", "Feed", "In"],
        [2, 1, "Test Opponent", "none", "Backhand", "Net"],
    ]
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows)
    path = tmp_path / "with_feed_point.xlsx"
    workbook.save(path)

    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(path, date="2026-08-06", opponent="Alex", result="W")
    record = load_pending(json_path)

    assert any("excluded as non-match activity" in note for note in record.import_notes)


def test_confirm_point_clears_a_flag_directly_without_an_api_call(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    target = record.sets[0].points[0]

    confirmed = pipeline.confirm_point(
        json_path,
        set_number=1,
        game_number=target.game_number,
        point_number=target.point_number,
        point_end_type="ace",
        point_won=True,
        net_approach=False,
    )

    updated = confirmed.sets[0].points[0]
    assert updated.point_end_type == "ace"
    assert updated.point_won is True
    assert updated.needs_review is False

    # Persisted to the same pending file.
    reloaded = load_pending(json_path)
    assert reloaded.sets[0].points[0].needs_review is False


def test_confirm_point_every_flagged_point_then_finalize_succeeds(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    for point in record.sets[0].points:
        pipeline.confirm_point(
            json_path,
            set_number=1,
            game_number=point.game_number,
            point_number=point.point_number,
            point_end_type="winner",
            point_won=True,
            net_approach=False,
        )

    match_id = pipeline.finalize(json_path)
    assert match_id == 1


def test_confirm_point_rejects_an_inconsistent_pair_and_leaves_the_point_flagged(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    target = record.sets[0].points[0]

    with pytest.raises(ConfirmationError, match="requires point_won"):
        pipeline.confirm_point(
            json_path,
            set_number=1,
            game_number=target.game_number,
            point_number=target.point_number,
            point_end_type="ace",
            point_won=False,
            net_approach=False,
        )

    reloaded = load_pending(json_path)
    assert reloaded.sets[0].points[0].needs_review is True


class _FakeSuggestionMessages:
    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        payload = json.dumps(
            {"point_end_type": "forced_error", "reasoning": "Deep shot.", "confidence": "medium"}
        )
        return FakeMessage(content=[FakeTextBlock(text=payload)])


class _FakeSuggestionClient:
    def __init__(self):
        self.messages = _FakeSuggestionMessages()


def test_suggest_annotates_reconstructed_points_without_touching_needs_review(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    client = _FakeSuggestionClient()

    record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert all(p.ai_suggested_point_end_type == "forced_error" for p in all_points)
    assert all(p.ai_suggestion_reasoning == "Deep shot." for p in all_points)
    assert all(p.needs_review for p in all_points)  # never cleared
    assert len(client.messages.calls) == len(all_points)

    # Suggestions persist to the same pending file.
    reloaded = load_pending(json_path)
    assert reloaded.sets[0].points[0].ai_suggested_point_end_type == "forced_error"


def test_suggest_raises_without_a_source_file_to_re_parse(import_config):
    pipeline = SwingVisionImportPipeline(import_config)
    record = MatchRecord(date="2026-08-06", opponent="Alex", result="W")
    json_path = save_pending(record, import_config.pending_dir)

    with pytest.raises(ValueError, match="source_file"):
        pipeline.suggest(_FakeSuggestionClient(), json_path)


def test_suggest_skips_direct_parse_points_with_no_shot_provenance(
    synthetic_xlsx, import_config
):
    # synthetic_xlsx has real Points-sheet rows, so ingest() takes the
    # direct-parse path - those points have no source_point_number and no
    # raw shots to reason over, so suggest() must leave them alone.
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    client = _FakeSuggestionClient()

    record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert all(p.source_point_number is None for p in all_points)
    assert all(p.ai_suggested_point_end_type is None for p in all_points)
    assert len(client.messages.calls) == 0


def test_ingest_raises_when_points_is_empty_and_settings_is_missing(tmp_path, import_config):
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )
    # A Settings sheet header with no data row at all -> raw.settings is None.
    workbook.create_sheet("Settings").append(["Host Team", "Guest Team"])
    workbook.create_sheet("Shots").append(["Point", "Shot", "Player", "Type", "Stroke", "Result"])
    path = tmp_path / "no_settings.xlsx"
    workbook.save(path)

    pipeline = SwingVisionImportPipeline(import_config)
    with pytest.raises(ValueError, match="Settings sheet"):
        pipeline.ingest(path, date="2026-08-06", opponent="Alex", result="W")


class _FlakySuggestionMessages:
    """Fails for the first point, succeeds for the rest."""

    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        if len(self.calls) == 1:
            return FakeMessage(content=[FakeTextBlock(text="not valid json")])
        payload = json.dumps(
            {"point_end_type": "winner", "reasoning": "Clean pass.", "confidence": "high"}
        )
        return FakeMessage(content=[FakeTextBlock(text=payload)])


class _FlakySuggestionClient:
    def __init__(self):
        self.messages = _FlakySuggestionMessages()


def test_suggest_continues_past_one_points_suggestion_failure(
    synthetic_non_pro_xlsx, import_config, caplog
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    client = _FlakySuggestionClient()

    with caplog.at_level(logging.ERROR):
        record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    succeeded = [p for p in all_points if p.ai_suggested_point_end_type == "winner"]
    failed = [p for p in all_points if p.ai_suggested_point_end_type is None]
    assert len(succeeded) == 1
    assert len(failed) == 1
    assert any("Suggestion failed" in r.message for r in caplog.records)


def _build_non_pro_xlsx(tmp_path, filename: str, shot_rows: list[list], **settings_kwargs):
    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows, **settings_kwargs)
    path = tmp_path / filename
    workbook.save(path)
    return path


def test_ingest_multi_part_merges_a_game_spanning_the_file_boundary(tmp_path, import_config):
    # 3 straight host points in part 1 - not enough alone to complete a
    # game (needs 4). 1 more host point in part 2, whose own Point counter
    # restarts from 1 - only merging correctly completes the game 4-0.
    part1 = _build_non_pro_xlsx(
        tmp_path,
        "part1.xlsx",
        [
            [1, 1, "Test Player", "first_serve", "Serve", "In"],
            [2, 1, "Test Player", "first_serve", "Serve", "In"],
            [3, 1, "Test Player", "first_serve", "Serve", "In"],
        ],
    )
    part2 = _build_non_pro_xlsx(
        tmp_path, "part2.xlsx", [[1, 1, "Test Player", "first_serve", "Serve", "In"]]
    )

    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest_multi_part(
        [part1, part2], date="2026-08-18", opponent="Real Opponent", result="W"
    )
    record = load_pending(json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert len(all_points) == 4
    assert record.source_files == [str(part1), str(part2)]
    assert record.sets[-1].games_won == 1
    assert record.sets[-1].games_lost == 0
    assert any("Merged from 2 files" in note for note in record.import_notes)


def test_ingest_multi_part_rejects_fewer_than_two_files(tmp_path, import_config):
    part1 = _build_non_pro_xlsx(
        tmp_path, "part1.xlsx", [[1, 1, "Test Player", "first_serve", "Serve", "In"]]
    )
    pipeline = SwingVisionImportPipeline(import_config)

    with pytest.raises(ValueError, match="at least 2 files"):
        pipeline.ingest_multi_part(
            [part1], date="2026-08-18", opponent="Real Opponent", result="W"
        )


def test_ingest_multi_part_rejects_files_that_disagree_on_host_name(tmp_path, import_config):
    part1 = _build_non_pro_xlsx(
        tmp_path,
        "part1.xlsx",
        [[1, 1, "Test Player", "first_serve", "Serve", "In"]],
        host_name="Rishi Gupta",
    )
    part2 = _build_non_pro_xlsx(
        tmp_path,
        "part2.xlsx",
        [[1, 1, "Someone Else", "first_serve", "Serve", "In"]],
        host_name="Someone Else",
    )
    pipeline = SwingVisionImportPipeline(import_config)

    with pytest.raises(ValueError, match="disagree"):
        pipeline.ingest_multi_part(
            [part1, part2], date="2026-08-18", opponent="Real Opponent", result="W"
        )


def test_ingest_multi_part_rejects_a_file_with_no_settings_sheet(tmp_path, import_config):
    part1 = _build_non_pro_xlsx(
        tmp_path, "part1.xlsx", [[1, 1, "Test Player", "first_serve", "Serve", "In"]]
    )

    workbook = Workbook()
    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )
    # A Settings sheet with no data row at all -> raw.settings is None
    # (matches test_ingest_raises_when_points_is_empty_and_settings_is_missing's
    # pattern - the sheet must exist or parse() itself raises an unrelated
    # KeyError for the missing sheet).
    workbook.create_sheet("Settings").append(["Host Team", "Guest Team"])
    workbook.create_sheet("Shots").append(["Point", "Shot", "Player", "Type", "Stroke", "Result"])
    part2 = tmp_path / "part2_no_settings.xlsx"
    workbook.save(part2)

    pipeline = SwingVisionImportPipeline(import_config)

    with pytest.raises(ValueError, match="Settings sheet"):
        pipeline.ingest_multi_part(
            [part1, part2], date="2026-08-18", opponent="Real Opponent", result="W"
        )


def test_ingest_multi_part_rejects_a_pro_export_in_the_mix(tmp_path, import_config, synthetic_xlsx):
    part2 = _build_non_pro_xlsx(
        tmp_path, "part2.xlsx", [[1, 1, "Test Player", "first_serve", "Serve", "In"]]
    )
    pipeline = SwingVisionImportPipeline(import_config)

    with pytest.raises(ValueError, match="Points sheet"):
        pipeline.ingest_multi_part(
            [synthetic_xlsx, part2], date="2026-08-18", opponent="Real Opponent", result="W"
        )


def test_suggest_re_merges_multi_part_files_using_source_point_number(tmp_path, import_config):
    part1 = _build_non_pro_xlsx(
        tmp_path,
        "part1.xlsx",
        [
            [1, 1, "Test Player", "first_serve", "Serve", "In"],
            [2, 1, "Test Player", "first_serve", "Serve", "In"],
            [3, 1, "Test Player", "first_serve", "Serve", "In"],
        ],
    )
    part2 = _build_non_pro_xlsx(
        tmp_path, "part2.xlsx", [[1, 1, "Test Player", "first_serve", "Serve", "In"]]
    )
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest_multi_part(
        [part1, part2], date="2026-08-18", opponent="Real Opponent", result="W"
    )
    client = _FakeSuggestionClient()

    record = pipeline.suggest(client, json_path)

    all_points = [p for s in record.sets for p in s.points]
    assert all(p.ai_suggested_point_end_type == "forced_error" for p in all_points)
    assert len(client.messages.calls) == len(all_points)


class _FakeResolutionMessages:
    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        payload = json.dumps(
            {
                "point_end_type": "winner",
                "point_won": True,
                "net_approach": False,
                "reasoning": "The human said it was a clean winner.",
            }
        )
        return FakeMessage(content=[FakeTextBlock(text=payload)])


class _FakeResolutionClient:
    def __init__(self):
        self.messages = _FakeResolutionMessages()


class _FlakyResolutionMessages:
    """Fails for the first point, succeeds for the rest."""

    def __init__(self):
        self.calls = []

    def create(self, **kwargs):
        self.calls.append(kwargs)
        if len(self.calls) == 1:
            return FakeMessage(content=[FakeTextBlock(text="not valid json")])
        payload = json.dumps(
            {"point_end_type": "winner", "point_won": True, "net_approach": False, "reasoning": "x"}
        )
        return FakeMessage(content=[FakeTextBlock(text=payload)])


class _FlakyResolutionClient:
    def __init__(self):
        self.messages = _FlakyResolutionMessages()


def test_resolve_parses_review_answers_without_touching_needs_review(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    record.sets[0].points[0].review_answer = "She was way out of position - clean winner."
    save_pending(record, import_config.pending_dir)
    client = _FakeResolutionClient()

    resolved = pipeline.resolve(client, json_path)

    answered_point = resolved.sets[0].points[0]
    assert answered_point.resolved_point_end_type == "winner"
    assert answered_point.resolved_point_won is True
    assert answered_point.resolution_reasoning == "The human said it was a clean winner."
    assert answered_point.needs_review is True  # never cleared by resolve()
    assert answered_point.point_end_type != "winner"  # real field untouched

    # The other point had no review_answer - left completely alone.
    untouched_point = resolved.sets[0].points[1]
    assert untouched_point.resolved_point_end_type is None
    assert len(client.messages.calls) == 1


def test_resolve_works_for_direct_parse_points_with_no_shot_context(
    synthetic_xlsx, import_config
):
    # synthetic_xlsx takes the direct-parse path - its points have no
    # source_point_number and no raw shots to enrich the prompt with, but
    # resolve() should still work from the review_answer alone.
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(synthetic_xlsx, date="2026-08-06", opponent="Alex", result="W")
    record = load_pending(json_path)
    flagged_point = next(p for s in record.sets for p in s.points if p.needs_review)
    flagged_point.review_answer = "That was actually a forced error, she was pulled wide."
    save_pending(record, import_config.pending_dir)
    client = _FakeResolutionClient()

    resolved = pipeline.resolve(client, json_path)

    answered = next(p for s in resolved.sets for p in s.points if p.review_answer)
    assert answered.resolved_point_end_type == "winner"
    assert answered.source_point_number is None


def test_resolve_continues_past_one_points_parse_failure(
    synthetic_non_pro_xlsx, import_config, caplog
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    for point in record.sets[0].points:
        point.review_answer = "Clean winner."
    save_pending(record, import_config.pending_dir)
    client = _FlakyResolutionClient()

    with caplog.at_level(logging.ERROR):
        resolved = pipeline.resolve(client, json_path)

    all_points = [p for s in resolved.sets for p in s.points]
    succeeded = [p for p in all_points if p.resolved_point_end_type == "winner"]
    failed = [p for p in all_points if p.resolved_point_end_type is None]
    assert len(succeeded) == 1
    assert len(failed) == 1
    assert any("Failed to parse review_answer" in r.message for r in caplog.records)


def test_apply_resolutions_applies_fields_and_clears_needs_review(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    for point in record.sets[0].points:
        point.review_answer = "Clean winner, no net play."
    save_pending(record, import_config.pending_dir)
    pipeline.resolve(_FakeResolutionClient(), json_path)

    applied = pipeline.apply_resolutions(json_path)

    all_points = [p for s in applied.sets for p in s.points]
    assert all(p.point_end_type == "winner" for p in all_points)
    assert all(p.point_won is True for p in all_points)
    assert all(p.needs_review is False for p in all_points)

    # Fully resolved now - finalize() should succeed instead of raising.
    match_id = pipeline.finalize(json_path)
    assert match_id == 1


def test_apply_resolutions_leaves_unresolved_points_untouched(
    synthetic_non_pro_xlsx, import_config
):
    pipeline = SwingVisionImportPipeline(import_config)
    json_path = pipeline.ingest(
        synthetic_non_pro_xlsx, date="2026-08-06", opponent="Alex", result="W"
    )
    record = load_pending(json_path)
    record.sets[0].points[0].review_answer = "Clean winner."
    save_pending(record, import_config.pending_dir)
    pipeline.resolve(_FakeResolutionClient(), json_path)

    applied = pipeline.apply_resolutions(json_path)

    untouched = applied.sets[0].points[1]
    assert untouched.needs_review is True
    assert untouched.resolved_point_end_type is None

    with pytest.raises(UnresolvedReviewError):
        pipeline.finalize(json_path)
