from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook

from swingvision_import.config import ImportConfig
from swingvision_import.db import get_connection
from swingvision_import.load import finalize_and_load
from swingvision_import.records import MatchRecord, PointRecord, SetRecord
from tests.swingvision_import.conftest import add_settings_and_shots_sheets
from webapp.app import create_app
from webapp.config import WebAppConfig


def _build_non_pro_workbook() -> Workbook:
    """Mirrors tests/swingvision_import/conftest.py's synthetic_non_pro_xlsx
    shape, but saved to an in-memory buffer instead of a tmp_path file, since
    the webapp needs upload bytes, not a filesystem path."""
    workbook = Workbook()

    sets_sheet = workbook.active
    sets_sheet.title = "Sets"
    sets_sheet.append(["Set", "Host Score", "Guest Score", "Set Winner"])
    sets_sheet.append([1, 6, 4, "host"])

    workbook.create_sheet("Games").append(["Game", "Set", "Server", "Game Winner"])
    workbook.create_sheet("Points").append(
        ["Point", "Game", "Set", "Match Server", "Point Winner", "Detail"]
    )

    # Point 1: host aces. Point 2 is a gap (no shots at all). Point 3: host
    # serves, guest nets a return - host wins.
    shot_rows = [
        [1, 1, "Test Player", "first_serve", "Serve", "In"],
        [3, 1, "Test Player", "first_serve", "Serve", "In"],
        [3, 2, "Test Opponent", "first_return", "Backhand", "Net"],
    ]
    add_settings_and_shots_sheets(workbook, shot_rows=shot_rows)
    return workbook


@pytest.fixture
def xlsx_bytes() -> bytes:
    buffer = io.BytesIO()
    _build_non_pro_workbook().save(buffer)
    return buffer.getvalue()


@pytest.fixture
def import_config(tmp_path: Path) -> ImportConfig:
    return ImportConfig(pending_dir=tmp_path / "pending", db_path=tmp_path / "rallyai.db")


@pytest.fixture
def webapp_config(tmp_path: Path) -> WebAppConfig:
    return WebAppConfig(uploads_dir=tmp_path / "uploads", media_dir=tmp_path / "media")


@pytest.fixture
def app(import_config: ImportConfig, webapp_config: WebAppConfig):
    flask_app = create_app(import_config, webapp_config)
    flask_app.config["TESTING"] = True
    return flask_app


@pytest.fixture
def client(app):
    return app.test_client()


@pytest.fixture
def finalized_match_id(import_config: ImportConfig) -> int:
    """Seeds a fully-reviewed, finalized match directly into the app's own
    db_path, so it's visible to the running app's routes exactly like a
    real match that went through review + finalize() would be."""
    connection = get_connection(import_config.db_path, import_config.schema_path)
    record = MatchRecord(
        date="2026-08-06",
        opponent="Alex",
        result="W",
        sets=[
            SetRecord(
                set_number=1,
                games_won=6,
                games_lost=4,
                points=[
                    PointRecord(1, 1, True, True, "ace"),
                    PointRecord(1, 2, False, False, "unforced_error"),
                ],
            ),
        ],
    )
    match_id = finalize_and_load(connection, record)
    connection.close()
    return match_id
